#!/usr/bin/env python3
"""
Excel Table Output Creator

Creates a formatted Excel file that preserves the original table structure
while providing the holistic combined view.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def create_formatted_excel():
    """Create Excel file with original format preserved"""
    
    print("🔄 Creating formatted Excel output...")
    
    # Read original data
    definitions = pd.read_excel('inputs/data.xlsx', sheet_name='definitions')
    monitors = pd.read_excel('inputs/data.xlsx', sheet_name='monitors')  
    conditions = pd.read_excel('inputs/data.xlsx', sheet_name='conditions')
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # 1. HOLISTIC VIEW SHEET - Main combined view
    print("✓ Creating Holistic View sheet...")
    ws_holistic = wb.create_sheet("Holistic_View")
    
    # Create holistic data structure
    holistic_data = []
    
    for _, def_row in definitions.iterrows():
        fdir_name = def_row['FDIRs']
        fdir_id = def_row['id']
        
        # Get monitors for this FDIR
        fdir_monitors = monitors[monitors['id'] == fdir_id]
        # Get conditions for this FDIR  
        fdir_conditions = conditions[conditions['id'] == fdir_id].dropna(subset=['condition_mons'])
        
        # Add header row for each FDIR
        holistic_data.append([fdir_name.upper(), fdir_id, '', '', '', '', ''])
        holistic_data.append(['MONITORS', '', '', '', '', '', ''])
        holistic_data.append(['Monitor Name', 'Threshold', '', 'CONDITIONS', '', '', ''])
        holistic_data.append(['Condition', 'Count', 'Response', '', '', '', ''])
        
        # Combine monitors and conditions in parallel columns
        max_rows = max(len(fdir_monitors), len(fdir_conditions))
        
        for i in range(max_rows):
            row = ['', '', '', '', '', '', '']
            
            # Add monitor data
            if i < len(fdir_monitors):
                mon_row = fdir_monitors.iloc[i]
                row[0] = mon_row['mons']
                row[1] = mon_row['thresholds']
            
            # Add condition data  
            if i < len(fdir_conditions):
                cond_row = fdir_conditions.iloc[i]
                row[3] = cond_row['condition_mons']
                row[4] = cond_row['counts'] 
                row[5] = cond_row['response']
            
            holistic_data.append(row)
        
        # Add separator
        holistic_data.append(['', '', '', '', '', '', ''])
    
    # Write holistic data
    headers = ['FDIR/Monitor', 'Threshold', '', 'Condition', 'Count', 'Response', '']
    ws_holistic.append(headers)
    
    for row in holistic_data:
        ws_holistic.append(row)
    
    # 2. ORIGINAL FORMAT SHEETS
    print("✓ Creating original format sheets...")
    
    # Definitions sheet (preserved format)
    ws_def = wb.create_sheet("Definitions")
    for r in dataframe_to_rows(definitions, index=False, header=True):
        ws_def.append(r)
    
    # Monitors sheet (preserved format) 
    ws_mon = wb.create_sheet("Monitors")
    for r in dataframe_to_rows(monitors, index=False, header=True):
        ws_mon.append(r)
    
    # Conditions sheet (preserved format)
    ws_cond = wb.create_sheet("Conditions") 
    for r in dataframe_to_rows(conditions, index=False, header=True):
        ws_cond.append(r)
    
    # 3. COMBINED FLAT VIEW SHEET
    print("✓ Creating combined flat view...")
    ws_flat = wb.create_sheet("Combined_Flat")
    
    # Create flat combined data preserving original columns
    flat_data = []
    
    for _, def_row in definitions.iterrows():
        fdir_name = def_row['FDIRs'] 
        fdir_id = def_row['id']
        
        # Get all monitors and conditions for this FDIR
        fdir_monitors = monitors[monitors['id'] == fdir_id]
        fdir_conditions = conditions[conditions['id'] == fdir_id].dropna(subset=['condition_mons'])
        
        # If no monitors or conditions, add one row
        if fdir_monitors.empty and fdir_conditions.empty:
            flat_data.append({
                'FDIRs': fdir_name,
                'id': fdir_id,
                'mons': '',
                'thresholds': '',
                'condition_mons': '',
                'counts': '',
                'response': ''
            })
        else:
            # Create combinations of monitors and conditions
            if fdir_monitors.empty:
                # Only conditions
                for _, cond_row in fdir_conditions.iterrows():
                    flat_data.append({
                        'FDIRs': fdir_name,
                        'id': fdir_id, 
                        'mons': '',
                        'thresholds': '',
                        'condition_mons': cond_row['condition_mons'],
                        'counts': cond_row['counts'],
                        'response': cond_row['response']
                    })
            elif fdir_conditions.empty:
                # Only monitors
                for _, mon_row in fdir_monitors.iterrows():
                    flat_data.append({
                        'FDIRs': fdir_name,
                        'id': fdir_id,
                        'mons': mon_row['mons'], 
                        'thresholds': mon_row['thresholds'],
                        'condition_mons': '',
                        'counts': '',
                        'response': ''
                    })
            else:
                # Both monitors and conditions - create all combinations
                max_items = max(len(fdir_monitors), len(fdir_conditions))
                
                for i in range(max_items):
                    row_data = {
                        'FDIRs': fdir_name,
                        'id': fdir_id,
                        'mons': '',
                        'thresholds': '', 
                        'condition_mons': '',
                        'counts': '',
                        'response': ''
                    }
                    
                    if i < len(fdir_monitors):
                        mon_row = fdir_monitors.iloc[i]
                        row_data['mons'] = mon_row['mons']
                        row_data['thresholds'] = mon_row['thresholds']
                    
                    if i < len(fdir_conditions):
                        cond_row = fdir_conditions.iloc[i]
                        row_data['condition_mons'] = cond_row['condition_mons']
                        row_data['counts'] = cond_row['counts']
                        row_data['response'] = cond_row['response']
                    
                    flat_data.append(row_data)
    
    # Write flat data
    flat_df = pd.DataFrame(flat_data)
    for r in dataframe_to_rows(flat_df, index=False, header=True):
        ws_flat.append(r)
    
    # 4. APPLY FORMATTING
    print("✓ Applying formatting...")
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format headers (first row)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders to all cells with data
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
    
    # Save the file
    output_file = 'outputs/combined_data_formatted.xlsx'
    wb.save(output_file)
    
    print(f"✅ Created formatted Excel file: {output_file}")
    
    # Summary
    print("\n📊 EXCEL FILE STRUCTURE:")
    print("="*50)
    print("📋 Holistic_View - Main combined view with side-by-side layout")
    print("📋 Definitions - Original definitions table (FDIRs, id)")
    print("📋 Monitors - Original monitors table (id, mons, thresholds)")  
    print("📋 Conditions - Original conditions table (id, condition_mons, counts, response)")
    print("📋 Combined_Flat - Flattened view with all original columns")
    
    return output_file

def main():
    """Main function"""
    print("📊 Excel Table Output Creator")
    print("="*50)
    
    try:
        output_file = create_formatted_excel()
        print(f"\n🎉 Success! Open '{output_file}' to view the formatted data.")
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()