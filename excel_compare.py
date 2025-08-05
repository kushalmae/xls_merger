#!/usr/bin/env python3
"""
Simple Excel Comparison Tool

Compares two Excel files and creates:
1. New Excel file with color-coded differences
2. CSV files for easy comparison

Usage: python excel_compare.py
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def compare_excel_files(file1_path, file2_path):
    """Compare two Excel files and return differences"""
    
    print(f"📊 Comparing Excel files:")
    print(f"  File 1: {file1_path}")
    print(f"  File 2: {file2_path}")
    
    # Read both Excel files
    try:
        df1 = pd.read_excel(file1_path)
        df2 = pd.read_excel(file2_path)
        print(f"✅ Successfully loaded both files")
        print(f"  File 1: {df1.shape[0]} rows, {df1.shape[1]} columns")
        print(f"  File 2: {df2.shape[0]} rows, {df2.shape[1]} columns")
        
    except Exception as e:
        print(f"❌ Error loading files: {e}")
        return None, None, None
    
    # Align both DataFrames to same size
    max_rows = max(len(df1), len(df2))
    max_cols = max(len(df1.columns), len(df2.columns))
    
    # Extend df1 if needed
    if len(df1) < max_rows:
        empty_rows = pd.DataFrame(index=range(len(df1), max_rows), columns=df1.columns)
        df1 = pd.concat([df1, empty_rows], ignore_index=True)
    
    # Extend df2 if needed  
    if len(df2) < max_rows:
        empty_rows = pd.DataFrame(index=range(len(df2), max_rows), columns=df2.columns)
        df2 = pd.concat([df2, empty_rows], ignore_index=True)
    
    # Align columns
    all_columns = list(set(df1.columns) | set(df2.columns))
    
    for col in all_columns:
        if col not in df1.columns:
            df1[col] = np.nan
        if col not in df2.columns:
            df2[col] = np.nan
    
    # Reorder columns to match
    df1 = df1.reindex(columns=all_columns)
    df2 = df2.reindex(columns=all_columns)
    
    # Create difference matrix
    differences = pd.DataFrame(index=df1.index, columns=df1.columns)
    
    for col in df1.columns:
        for row in df1.index:
            val1 = df1.loc[row, col]
            val2 = df2.loc[row, col]
            
            # Handle NaN comparisons
            if pd.isna(val1) and pd.isna(val2):
                differences.loc[row, col] = 'SAME'
            elif pd.isna(val1) or pd.isna(val2):
                differences.loc[row, col] = 'DIFFERENT'
            elif str(val1) == str(val2):
                differences.loc[row, col] = 'SAME'
            else:
                differences.loc[row, col] = 'DIFFERENT'
    
    print(f"✅ Comparison completed")
    
    # Count differences
    total_cells = differences.size
    different_cells = (differences == 'DIFFERENT').sum().sum()
    same_cells = total_cells - different_cells
    
    print(f"📈 Results:")
    print(f"  Total cells compared: {total_cells}")
    print(f"  Same cells: {same_cells}")
    print(f"  Different cells: {different_cells}")
    print(f"  Difference percentage: {(different_cells/total_cells*100):.1f}%")
    
    return df1, df2, differences

def create_comparison_excel(df1, df2, differences, output_file):
    """Create Excel file maintaining original layout with color-coded differences"""
    
    print(f"📝 Creating comparison Excel with original layout: {output_file}")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define colors - subtle highlighting to preserve readability
    same_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White (no highlight)
    diff_fill = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")  # Very light red
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Light blue
    header_font = Font(bold=True)
    
    # Sheet 1: Original File 1 Layout with highlighting
    ws_file1 = wb.create_sheet("Original_File1")
    
    # Add headers exactly as in original
    for col, header in enumerate(df1.columns, 1):
        cell = ws_file1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Add data maintaining exact original format
    for row in range(len(df1)):
        for col, col_name in enumerate(df1.columns, 1):
            value = df1.loc[row, col_name]
            
            # Preserve original value format (including NaN/empty cells)
            if pd.notna(value):
                cell_value = value
            else:
                cell_value = ""
            
            cell = ws_file1.cell(row=row+2, column=col, value=cell_value)
            
            # Only highlight differences with subtle color
            if differences.loc[row, col_name] == 'DIFFERENT':
                cell.fill = diff_fill
            # No highlighting for same values - keep clean look
    
    # Sheet 2: Original File 2 Layout with highlighting  
    ws_file2 = wb.create_sheet("Original_File2")
    
    # Add headers exactly as in original
    for col, header in enumerate(df2.columns, 1):
        cell = ws_file2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Add data maintaining exact original format
    for row in range(len(df2)):
        for col, col_name in enumerate(df2.columns, 1):
            value = df2.loc[row, col_name]
            
            # Preserve original value format
            if pd.notna(value):
                cell_value = value
            else:
                cell_value = ""
            
            cell = ws_file2.cell(row=row+2, column=col, value=cell_value)
            
            # Only highlight differences with subtle color
            if differences.loc[row, col_name] == 'DIFFERENT':
                cell.fill = diff_fill
            # No highlighting for same values - keep clean look
    
    # Sheet 3: Side-by-side comparison (only if there are differences)
    different_cells = (differences == 'DIFFERENT').sum().sum()
    if different_cells > 0:
        ws_comparison = wb.create_sheet("Differences_Summary")
        
        # Create headers
        headers = ['Row', 'Column', 'File 1 Value', 'File 2 Value']
        for col, header in enumerate(headers, 1):
            cell = ws_comparison.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Add only different values
        row_num = 2
        for df_row in range(len(df1)):
            for col_name in df1.columns:
                if differences.loc[df_row, col_name] == 'DIFFERENT':
                    val1 = df1.loc[df_row, col_name]
                    val2 = df2.loc[df_row, col_name]
                    
                    ws_comparison.cell(row=row_num, column=1, value=df_row + 1)
                    ws_comparison.cell(row=row_num, column=2, value=col_name)
                    ws_comparison.cell(row=row_num, column=3, value=str(val1) if pd.notna(val1) else "")
                    ws_comparison.cell(row=row_num, column=4, value=str(val2) if pd.notna(val2) else "")
                    
                    # Highlight difference rows
                    for col in range(1, 5):
                        ws_comparison.cell(row=row_num, column=col).fill = diff_fill
                    
                    row_num += 1
    
    # Apply text wrapping and formatting to preserve newlines
    from openpyxl.styles import Alignment
    
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    for ws in wb.worksheets:
        # Apply text wrapping to all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = wrap_alignment
        
        # Auto-adjust column widths considering newlines
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        # Handle newlines in length calculation
                        lines = str(cell.value).split('\n')
                        cell_length = max(len(line) for line in lines)
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set reasonable width that preserves readability
            adjusted_width = max(min(max_length + 3, 60), 15)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(output_file)
    print(f"✅ Excel comparison saved with original layout preserved: {output_file}")

def create_comparison_csvs(df1, df2, base_name):
    """Create CSV files for comparison"""
    
    print(f"📄 Creating CSV files...")
    
    # Save individual files
    csv1 = f"{base_name}_file1.csv"
    csv2 = f"{base_name}_file2.csv"
    
    df1.to_csv(csv1, index=False)
    df2.to_csv(csv2, index=False)
    
    print(f"✅ CSV files created:")
    print(f"  {csv1}")
    print(f"  {csv2}")
    
    return csv1, csv2

def main():
    """Main function"""
    print("🔍 Excel Comparison Tool")
    print("="*40)
    
    # Use the specific files requested
    file1 = "inputs/response_text.xlsx"
    file2 = "inputs/response_text2.xlsx"
    
    # Check if files exist
    if not os.path.exists(file1):
        print(f"❌ File not found: {file1}")
        return
    
    if not os.path.exists(file2):
        print(f"❌ File not found: {file2}")
        return
    
    # Compare files
    df1, df2, differences = compare_excel_files(file1, file2)
    
    if df1 is None:
        return
    
    # Create outputs directory
    os.makedirs("outputs", exist_ok=True)
    
    # Create comparison Excel with file locking handling
    from datetime import datetime
    
    base_excel = "outputs/excel_comparison.xlsx"
    excel_output = base_excel
    
    # If file exists and is locked, create a new version
    try:
        # Try to open for writing to check if locked
        with open(excel_output, 'a'):
            pass
    except PermissionError:
        # File is locked, create new version with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_output = f'outputs/excel_comparison_{timestamp}.xlsx'
        print(f"⚠️  Original file is open, creating: {excel_output}")
    
    # Create comparison Excel
    create_comparison_excel(df1, df2, differences, excel_output)
    
    # Create comparison CSVs
    csv_base = "outputs/comparison"
    create_comparison_csvs(df1, df2, csv_base)
    
    print("\n🎉 Comparison completed!")
    print("📋 Files created:")
    print(f"  📊 {excel_output} - Color-coded comparison")
    print(f"  📄 {csv_base}_file1.csv - First file as CSV")
    print(f"  📄 {csv_base}_file2.csv - Second file as CSV")
    
    print("\n📊 Excel sheets:")
    print("  📋 Original_File1 - First file in original layout with differences highlighted")
    print("  📋 Original_File2 - Second file in original layout with differences highlighted")
    
    # Check if there are differences to show summary sheet
    different_cells = (differences == 'DIFFERENT').sum().sum()
    if different_cells > 0:
        print("  🔍 Differences_Summary - Only the different cells listed")
    else:
        print("  ✅ No differences found - files are identical!")
    
    print("\n🎨 Color Legend:")
    print("  ⬜ White = Same values (no highlighting)")
    print("  🔴 Light Red = Different values")
    
    if different_cells == 0:
        print("\n🎯 Result: Files are identical - no differences found!")
    else:
        print(f"\n⚠️  Found {different_cells} different cells - check highlighted areas!")

if __name__ == "__main__":
    main()